import uuid
import asyncio
import logging
import os
import sys
import re
import random
import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse

from app.memory_store import (
    init_store,
    create_document,
    get_document,
    create_job,
    update_job,
    get_job,
    list_jobs as store_list_jobs,
    upsert_party,
    get_party as store_get_party,
    create_invoice,
    get_invoice as store_get_invoice,
    list_invoices as store_list_invoices,
    update_invoice as store_update_invoice,
    delete_invoice as store_delete_invoice,
    dumps,
    loads,
    parties_store,
)
from app.schemas import (
    JobStatus,
    Party,
    InvoiceDetail,
    InvoiceListItem,
    RequestRescanRequest,
    UpdateInvoiceRequest,
    UploadResponse,
    UploadJobResponse,
)

BASE_DIR = Path(__file__).resolve().parents[1]
STORAGE_DIR = BASE_DIR / "storage"
STORAGE_DIR.mkdir(parents=True, exist_ok=True)

# Ensure repo root is importable (so we can import sibling invoice_extractor/)
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


def _load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


_load_env_file(REPO_ROOT / "invoice_extractor" / ".env")
API_KEY = os.getenv("GEMINI_API_KEY")

# Import extractor library from sibling folder (after sys.path fix)
from invoice_extractor.lib import INVOICE_FIELDS, extract_from_pdf_bytes  # noqa: E402

app = FastAPI(title="Invoice Backend (Demo - No Auth)", version="0.2.0")

logger = logging.getLogger("invoice-backend")
if not logger.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")

def _env_int(name: str, default: int) -> int:
    try:
        v = int(str(os.getenv(name, "")).strip() or default)
        return v
    except Exception:
        return default


# Performance tuning (safe defaults)
EXTRACT_DPI = max(100, min(_env_int("INVOICE_EXTRACT_DPI", 200), 400))
RETRY_DPI = max(EXTRACT_DPI, min(_env_int("INVOICE_RETRY_DPI", 300), 600))
MAX_PAGE_CONCURRENCY = max(1, min(_env_int("INVOICE_PAGE_CONCURRENCY", 3), 3))
PAGE_TIMEOUT_S = max(30, min(_env_int("INVOICE_PAGE_TIMEOUT_S", 180), 900))
BIG_PDF_PAGES = max(1, min(_env_int("INVOICE_BIG_PDF_PAGES", 10), 500))
BIG_PDF_BYTES = max(1024 * 1024, min(_env_int("INVOICE_BIG_PDF_BYTES", 8 * 1024 * 1024), 200 * 1024 * 1024))
VLM_BATCH_SIZE = max(1, min(_env_int("INVOICE_VLM_BATCH_SIZE", 3), 6))
VLM_IMAGE_FORMAT = (os.getenv("INVOICE_VLM_IMAGE_FORMAT", "jpeg") or "jpeg").strip().lower()

# CORS - allow all origins for demo deployment
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for demo
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _startup():
    init_store()
    logger.info("Demo mode: Using in-memory storage (no persistence)")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


async def _extract_pages_with_retry(
    *,
    pdf_bytes: bytes,
    only_page: int | None = None,
    page_numbers: list[int] | None = None,
    dpi: int = EXTRACT_DPI,
    retry_dpi: int = RETRY_DPI,
):
    """
    Run extraction off the event loop (thread) and retry transient Gemini failures.
    """
    if not API_KEY:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured on backend.")

    backoff_s = [1.0, 2.0, 4.0, 8.0, 16.0]
    last_err: Exception | None = None

    for attempt in range(len(backoff_s) + 1):
        try:
            return await asyncio.wait_for(
                asyncio.to_thread(
                    extract_from_pdf_bytes,
                    pdf_bytes,
                    api_key=API_KEY,
                    dpi=dpi,
                    model_name="gemini-2.0-flash-exp",
                    only_page=only_page,
                    page_numbers=page_numbers,
                    render_retry=True,
                    retry_dpi=retry_dpi,
                    batch_size=1 if only_page is not None else VLM_BATCH_SIZE,
                    image_format=VLM_IMAGE_FORMAT,
                ),
                timeout=PAGE_TIMEOUT_S,
            )
        except Exception as e:
            last_err = e
            msg = str(e) or ""
            msg_u = msg.upper()
            msg_l = msg.lower()

            is_503 = ("503" in msg) or ("UNAVAILABLE" in msg_u)
            is_transient_net = any(
                s in msg_l
                for s in [
                    "getaddrinfo failed",
                    "name or service not known",
                    "temporary failure in name resolution",
                    "errno 11001",
                    "server disconnected",
                    "unexpected_eof_while_reading",
                    "eof occurred in violation of protocol",
                    "connection reset",
                    "connection aborted",
                    "timed out",
                ]
            )
            is_timeout = isinstance(e, asyncio.TimeoutError) or ("timeout" in msg_l)
            is_retryable = is_503 or is_transient_net or is_timeout

            if is_retryable and attempt < len(backoff_s):
                base = backoff_s[attempt]
                delay = base * random.uniform(0.8, 1.3)
                logger.warning(
                    "Extraction retryable error (attempt %s), retrying in %.1fs (page=%s): %s",
                    attempt + 1,
                    delay,
                    only_page,
                    msg,
                )
                await asyncio.sleep(delay)
                continue

            if is_503:
                logger.exception("Gemini unavailable after retries")
                raise HTTPException(
                    status_code=503,
                    detail="Gemini service temporarily unavailable. Please retry in a minute.",
                )

            logger.exception("Extraction failed")
            raise

    raise last_err  # type: ignore[misc]


def _job_to_model(job: dict) -> JobStatus:
    return JobStatus(
        id=job["id"],
        document_id=job["document_id"],
        filename=job.get("filename"),
        status=job["status"],
        total_pages=job.get("total_pages"),
        processed_pages=int(job.get("processed_pages") or 0),
        message=job.get("message"),
        error=job.get("error"),
        invoice_ids=job.get("invoice_ids", []),
        has_low_readability=bool(job.get("has_low_readability", False)),
        created_at=job.get("created_at"),
        updated_at=job.get("updated_at"),
    )


def _is_low_readability(p) -> bool:
    """
    True only when the page has actual readability/vision issues.
    """
    reasons = set((getattr(p, "reasons", None) or []) if isinstance(getattr(p, "reasons", None), list) else [])
    if any(r.startswith("json_parse_failed") for r in reasons):
        return True
    if "model_flagged_unreadable" in reasons:
        return True

    diag = getattr(p, "field_diagnostics", None) or {}
    if isinstance(diag, dict):
        critical = {"Invoice_No", "Invoice_Date", "Net_Amount"}
        for f in critical:
            v = diag.get(f)
            if not isinstance(v, dict):
                continue
            status = str(v.get("status") or "").lower()
            if status not in {"unreadable", "blurry", "faded", "cut_off"}:
                continue
            if f == "Invoice_Date":
                reason = str(v.get("reason") or "").lower()
                if any(s in reason for s in ["format", "invalid day", "invalid month", "ambiguous date"]):
                    continue
            return True

    unreadable_fields = set(getattr(p, "unreadable_fields", None) or [])
    if bool(getattr(p, "needs_rescan", False)) and (unreadable_fields & {"Invoice_No", "Invoice_Date", "Net_Amount"}):
        return True
    return False


def _insert_invoice_for_page(*, doc_id: str, p) -> tuple[str, bool, bool]:
    """
    Insert a single invoice for a page extraction.
    Returns (invoice_id, needs_review, low_readability)
    """
    inv_id = uuid.uuid4().hex
    extracted = dict(p.data)

    system_conf = p.system_confidence if p.system_confidence is not None else 1.0
    diag = p.field_diagnostics or {}
    needs_audit = any(
        isinstance(v, dict) and v.get("status") in {"ambiguous"}
        for v in diag.values()
    )
    low_readability = _is_low_readability(p)
    needs_review = bool(p.needs_rescan) or bool(needs_audit) or (system_conf < 0.85)
    status = "needs-review" if needs_review else "auto-extracted"

    supplier_party_id = upsert_party(
        party_type="supplier",
        name=extracted.get("Supplier_Name"),
        ntn=extracted.get("Supplier_NTN"),
        gst_no=extracted.get("Supplier_GST_No"),
        registration_no=extracted.get("Supplier_Registration_No"),
    )
    buyer_party_id = upsert_party(
        party_type="buyer",
        name=extracted.get("Buyer_Name"),
        ntn=extracted.get("Buyer_NTN"),
        gst_no=extracted.get("Buyer_GST_No"),
        registration_no=extracted.get("Buyer_Registration_No"),
    )

    create_invoice(
        invoice_id=inv_id,
        document_id=doc_id,
        page_no=int(getattr(p, "page_no", 0) or 0),
        supplier_party_id=supplier_party_id,
        buyer_party_id=buyer_party_id,
        extracted_data=extracted,
        status=status,
        needs_rescan=p.needs_rescan,
        unreadable_fields=p.unreadable_fields or [],
        reasons=p.reasons or [],
        model_avg_confidence=getattr(p, "avg_field_confidence", None),
        system_confidence=getattr(p, "system_confidence", None),
        system_reasons=getattr(p, "system_reasons", []) or [],
        field_diagnostics=getattr(p, "field_diagnostics", {}) or {},
    )

    return inv_id, needs_review, low_readability


async def _run_document_job(*, job_id: str, doc_id: str, stored_path: Path) -> None:
    """
    Background job that extracts pages and writes invoices incrementally.
    """
    try:
        update_job(job_id=job_id, status="running", message="Reading PDF...")
        pdf_bytes = stored_path.read_bytes()
        import fitz

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_pages = len(doc)
        doc.close()

        update_job(job_id=job_id, total_pages=total_pages, processed_pages=0, message="Extracting pages...")

        invoice_ids: list[str] = []
        any_low_readability = False
        processed_pages = 0

        update_job(job_id=job_id, message=f"Extracting pages... (0/{total_pages})")

        # Extract first page
        first_pages = await _extract_pages_with_retry(
            pdf_bytes=pdf_bytes,
            only_page=1,
            dpi=EXTRACT_DPI,
            retry_dpi=RETRY_DPI,
        )
        first = first_pages[0] if first_pages else None
        if first is None:
            raise RuntimeError("Extraction returned no pages for page 1")

        inv_id, _needs_review, low_readability = _insert_invoice_for_page(doc_id=doc_id, p=first)
        invoice_ids.append(inv_id)
        any_low_readability = any_low_readability or low_readability
        processed_pages = 1
        update_job(
            job_id=job_id,
            processed_pages=processed_pages,
            invoice_ids=invoice_ids,
            has_low_readability=any_low_readability,
            message=f"Extracting pages... ({processed_pages}/{total_pages})",
        )

        effective_concurrency = MAX_PAGE_CONCURRENCY
        if (total_pages >= BIG_PDF_PAGES or len(pdf_bytes) >= BIG_PDF_BYTES) and any_low_readability:
            effective_concurrency = 1
            update_job(
                job_id=job_id,
                message=f"Low-readability big PDF detected; switching to sequential processing (1/{total_pages})",
            )

        sem = asyncio.Semaphore(effective_concurrency)
        state_lock = asyncio.Lock()

        def _chunks(seq: list[int], n: int) -> list[list[int]]:
            n = max(1, int(n or 1))
            return [seq[i : i + n] for i in range(0, len(seq), n)]

        async def _process_batch(page_nums_1based: list[int]) -> None:
            nonlocal any_low_readability, processed_pages
            if not page_nums_1based:
                return
            async with sem:
                pages = await _extract_pages_with_retry(
                    pdf_bytes=pdf_bytes,
                    page_numbers=page_nums_1based,
                    dpi=EXTRACT_DPI,
                    retry_dpi=RETRY_DPI,
                )
                if not pages:
                    raise RuntimeError(f"Extraction returned no pages for batch {page_nums_1based}")

                inv_ids: list[str] = []
                batch_low = False
                for p in pages:
                    inv_id, _needs_review, low_readability = _insert_invoice_for_page(doc_id=doc_id, p=p)
                    inv_ids.append(inv_id)
                    batch_low = batch_low or low_readability

                async with state_lock:
                    invoice_ids.extend(inv_ids)
                    any_low_readability = any_low_readability or batch_low
                    processed_pages += len(inv_ids)
                    update_job(
                        job_id=job_id,
                        processed_pages=processed_pages,
                        invoice_ids=invoice_ids,
                        has_low_readability=any_low_readability,
                        message=f"Extracting pages... ({processed_pages}/{total_pages})",
                    )

        remaining = list(range(2, total_pages + 1))
        batches = _chunks(remaining, VLM_BATCH_SIZE)
        tasks = [asyncio.create_task(_process_batch(b)) for b in batches]
        try:
            await asyncio.gather(*tasks)
        except Exception:
            for t in tasks:
                t.cancel()
            raise

        update_job(
            job_id=job_id,
            status="completed",
            message="Completed",
            processed_pages=total_pages,
            invoice_ids=invoice_ids,
            has_low_readability=any_low_readability,
        )
        logger.info("Job completed job_id=%s document_id=%s pages=%s", job_id, doc_id, total_pages)
    except Exception as e:
        logger.exception("Job failed job_id=%s document_id=%s", job_id, doc_id)
        update_job(job_id=job_id, status="failed", error=str(e), message="Failed")


@app.get("/api/jobs/{job_id}", response_model=JobStatus)
def get_job_endpoint(job_id: str):
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")
    return _job_to_model(job)


@app.get("/api/jobs", response_model=list[JobStatus])
def list_jobs_endpoint(limit: int = 50):
    limit = max(1, min(int(limit or 50), 200))
    jobs = store_list_jobs(limit)
    return [_job_to_model(j) for j in jobs]


@app.post("/api/documents", response_model=UploadJobResponse)
async def upload_document(file: UploadFile = File(...)):
    if file.content_type not in ("application/pdf", "application/x-pdf", "application/acrobat", "applications/vnd.pdf"):
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Only PDF uploads are supported.")

    if not API_KEY:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured on backend.")

    doc_id = uuid.uuid4().hex
    stored_path = STORAGE_DIR / f"{doc_id}.pdf"

    pdf_bytes = await file.read()
    stored_path.write_bytes(pdf_bytes)

    create_document(doc_id, file.filename or "upload.pdf", str(stored_path))

    job_id = uuid.uuid4().hex
    create_job(job_id, doc_id)

    logger.info("Enqueued job job_id=%s document_id=%s filename=%s", job_id, doc_id, file.filename)
    asyncio.create_task(_run_document_job(job_id=job_id, doc_id=doc_id, stored_path=stored_path))

    return UploadJobResponse(job_id=job_id, document_id=doc_id)


@app.get("/api/documents/{document_id}/file")
def get_document_file(document_id: str):
    doc = get_document(document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")
    
    path = doc["stored_path"]
    name = doc["filename"]

    headers = {"Content-Disposition": f'inline; filename="{name}"'}
    return FileResponse(path, media_type="application/pdf", headers=headers)


def _invoice_to_list_item(inv: dict) -> InvoiceListItem:
    """Convert invoice dict to InvoiceListItem"""
    extracted = loads(inv["extracted_json"]) or {}
    edited = loads(inv["edited_json"]) or {}
    current = edited or extracted
    
    return InvoiceListItem(
        id=inv["id"],
        document_id=inv["document_id"],
        page_no=inv["page_no"],
        supplier_party_id=inv.get("supplier_party_id"),
        buyer_party_id=inv.get("buyer_party_id"),
        status=inv["status"],
        needs_rescan=inv["needs_rescan"],
        unreadable_fields=loads(inv["unreadable_fields_json"]) or [],
        reasons=loads(inv["reasons_json"]) or [],
        extracted=extracted,
        current=current,
        model_avg_confidence=inv.get("model_avg_confidence"),
        system_confidence=inv.get("system_confidence"),
        system_reasons=loads(inv.get("system_reasons_json")) or [],
        field_diagnostics=loads(inv.get("field_diagnostics_json")) or {},
    )


@app.get("/api/invoices", response_model=list[InvoiceListItem])
def list_invoices_endpoint(include_history: bool = False):
    invoices = store_list_invoices()
    return [_invoice_to_list_item(inv) for inv in invoices]


@app.get("/api/invoices/export.xlsx")
def export_invoices_xlsx(include_history: bool = False):
    """
    Backend-generated Excel export of invoices.
    """
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export dependency missing (openpyxl). {e}")

    invoices = store_list_invoices()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice_Date",
        "Invoice_No",
        "Supplier_Name",
        "Supplier_NTN",
        "Supplier_GST_No",
        "Supplier_Registration_No",
        "Buyer_Name",
        "Buyer_NTN",
        "Buyer_GST_No",
        "Buyer_Registration_No",
        "Exclusive_Value",
        "GST_Sales_Tax",
        "Inclusive_Value",
        "Advance_Tax",
        "Net_Amount",
        "Return",
        "Discount",
        "Incentive",
        "Location",
        "GRN",
    ]
    ws.append(headers)

    for inv in invoices:
        extracted = loads(inv["extracted_json"]) or {}
        edited = loads(inv["edited_json"]) or {}
        current = edited or extracted
        ws.append(
            [
                current.get("Invoice_Date"),
                current.get("Invoice_No"),
                current.get("Supplier_Name"),
                current.get("Supplier_NTN"),
                current.get("Supplier_GST_No"),
                current.get("Supplier_Registration_No"),
                current.get("Buyer_Name"),
                current.get("Buyer_NTN"),
                current.get("Buyer_GST_No"),
                current.get("Buyer_Registration_No"),
                current.get("Exclusive_Value"),
                current.get("GST_Sales_Tax"),
                current.get("Inclusive_Value"),
                current.get("Advance_Tax"),
                current.get("Net_Amount"),
                current.get("Return"),
                current.get("Discount"),
                current.get("Incentive"),
                current.get("Location"),
                current.get("GRN"),
            ]
        )

    ws.freeze_panes = "A2"
    try:
        for col_idx, col_name in enumerate(headers, start=1):
            max_len = len(col_name)
            for row_idx in range(2, min(ws.max_row, 250) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 45)
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"invoices_{ts}.xlsx"
    headers_out = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_out,
    )


@app.get("/api/ai-review", response_model=list[InvoiceListItem])
def list_ai_review(include_history: bool = False):
    invoices = store_list_invoices()
    # Filter for needs-review or needs_rescan
    filtered = [inv for inv in invoices if inv["needs_rescan"] or inv["status"] == "needs-review"]
    return [_invoice_to_list_item(inv) for inv in filtered]


@app.get("/api/invoices/{invoice_id}", response_model=InvoiceDetail)
def get_invoice_endpoint(invoice_id: str):
    inv = store_get_invoice(invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")
    
    extracted = loads(inv["extracted_json"]) or {}
    edited = loads(inv["edited_json"]) or {}
    current = edited or extracted
    document_id = inv["document_id"]
    page_no = int(inv["page_no"])

    document_url = f"/api/documents/{document_id}/file#page={page_no}"

    return InvoiceDetail(
        id=inv["id"],
        document_id=document_id,
        page_no=page_no,
        supplier_party_id=inv.get("supplier_party_id"),
        buyer_party_id=inv.get("buyer_party_id"),
        status=inv["status"],
        needs_rescan=inv["needs_rescan"],
        unreadable_fields=loads(inv["unreadable_fields_json"]) or [],
        reasons=loads(inv["reasons_json"]) or [],
        extracted=extracted,
        edited=edited or extracted,
        current=current,
        model_avg_confidence=inv.get("model_avg_confidence"),
        system_confidence=inv.get("system_confidence"),
        system_reasons=loads(inv.get("system_reasons_json")) or [],
        field_diagnostics=loads(inv.get("field_diagnostics_json")) or {},
        document_url=document_url,
    )


@app.get("/api/parties", response_model=list[Party])
def list_parties(party_type: str | None = None):
    parties = list(parties_store.values())
    
    if party_type:
        parties = [p for p in parties if p["type"] == party_type]
    
    # Sort by name
    parties.sort(key=lambda p: (p.get("name_raw") or "").lower())
    
    return [
        Party(
            id=p["id"],
            type=p["type"],
            name=p.get("name_raw"),
            ntn=p.get("ntn_raw"),
            gst_no=p.get("gst_raw"),
            registration_no=p.get("registration_raw"),
            created_at=p.get("created_at"),
            updated_at=p.get("updated_at"),
        )
        for p in parties
    ]


@app.get("/api/parties/{party_id}", response_model=Party)
def get_party_endpoint(party_id: str):
    p = store_get_party(party_id)
    if not p:
        raise HTTPException(status_code=404, detail="Party not found.")
    
    return Party(
        id=p["id"],
        type=p["type"],
        name=p.get("name_raw"),
        ntn=p.get("ntn_raw"),
        gst_no=p.get("gst_raw"),
        registration_no=p.get("registration_raw"),
        created_at=p.get("created_at"),
        updated_at=p.get("updated_at"),
    )


@app.put("/api/invoices/{invoice_id}", response_model=InvoiceDetail)
def update_invoice_endpoint(invoice_id: str, payload: UpdateInvoiceRequest):
    inv = store_get_invoice(invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    extracted = loads(inv["extracted_json"]) or {}
    edited = loads(inv["edited_json"]) or extracted

    if payload.edited is not None and isinstance(payload.edited, dict):
        next_edited: dict[str, Any] = {k: payload.edited.get(k) for k in INVOICE_FIELDS}
        edited = next_edited

    status = payload.status or inv["status"]

    # Approving clears rescan flag
    if status == "approved":
        inv["needs_rescan"] = False

    store_update_invoice(invoice_id, edited_data=edited, status=status)

    return get_invoice_endpoint(invoice_id)


@app.post("/api/invoices/{invoice_id}/request-rescan", response_model=InvoiceDetail)
def request_rescan(invoice_id: str, payload: RequestRescanRequest):
    inv = store_get_invoice(invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    existing_unreadable = loads(inv["unreadable_fields_json"]) or []
    existing_reasons = loads(inv["reasons_json"]) or []

    unreadable = sorted(set(existing_unreadable + (payload.unreadable_fields or [])))
    reasons = sorted(set(existing_reasons + (payload.reasons or []) + ["user_requested_rescan"]))

    inv["needs_rescan"] = True
    inv["status"] = "needs-review"
    inv["unreadable_fields_json"] = dumps(unreadable)
    inv["reasons_json"] = dumps(reasons)
    inv["updated_at"] = _now_iso()

    return get_invoice_endpoint(invoice_id)


@app.post("/api/invoices/{invoice_id}/reupload", response_model=InvoiceDetail)
async def reupload_invoice(invoice_id: str, file: UploadFile = File(...)):
    """
    Replace the document backing an invoice and re-run extraction.
    """
    if not API_KEY:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured on backend.")

    if file.content_type not in ("application/pdf", "application/x-pdf", "application/acrobat", "applications/vnd.pdf"):
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Only PDF uploads are supported.")

    inv = store_get_invoice(invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    pdf_bytes = await file.read()

    # Store new document
    new_doc_id = uuid.uuid4().hex
    stored_path = STORAGE_DIR / f"{new_doc_id}.pdf"
    stored_path.write_bytes(pdf_bytes)

    create_document(new_doc_id, file.filename or "reupload.pdf", str(stored_path))

    # Re-extract first page
    logger.info("Starting reupload extraction for invoice_id=%s new_doc_id=%s bytes=%s", invoice_id, new_doc_id, len(pdf_bytes))
    pages = await _extract_pages_with_retry(pdf_bytes=pdf_bytes, only_page=1)
    logger.info("Finished reupload extraction for invoice_id=%s new_doc_id=%s pages=%s", invoice_id, new_doc_id, len(pages))
    
    if not pages:
        raise HTTPException(status_code=400, detail="Uploaded PDF has no pages.")
    
    p = pages[0]
    extracted = dict(p.data)
    status = "needs-review" if p.needs_rescan else "auto-extracted"

    inv["document_id"] = new_doc_id
    inv["page_no"] = 1
    inv["extracted_json"] = dumps(extracted)
    inv["edited_json"] = None
    inv["status"] = status
    inv["needs_rescan"] = p.needs_rescan
    inv["unreadable_fields_json"] = dumps(p.unreadable_fields)
    inv["reasons_json"] = dumps(sorted(set((p.reasons or []) + ["reuploaded_and_reprocessed"])))
    inv["updated_at"] = _now_iso()

    return get_invoice_endpoint(invoice_id)
